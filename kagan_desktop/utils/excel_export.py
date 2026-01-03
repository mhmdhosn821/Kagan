"""
خروجی Excel از گزارشات
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any
from datetime import datetime
import os


class ExcelExporter:
    """کلاس صادرات داده به Excel"""
    
    def __init__(self):
        self.workbook = None
        self.worksheet = None
        
    def create_workbook(self, sheet_name: str = "Sheet1"):
        """ایجاد workbook جدید"""
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = sheet_name
        
    def set_headers(self, headers: List[str], rtl: bool = True):
        """تنظیم هدرهای جدول"""
        if not self.worksheet:
            return
        
        # استایل هدر
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # نوشتن هدرها
        for col_num, header in enumerate(headers, 1):
            cell = self.worksheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # تنظیم RTL
        if rtl:
            self.worksheet.sheet_view.rightToLeft = True
    
    def add_rows(self, data: List[List[Any]]):
        """افزودن ردیف‌های داده"""
        if not self.worksheet:
            return
        
        # استایل سلول‌ها
        cell_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # افزودن داده‌ها
        start_row = self.worksheet.max_row + 1
        
        for row_num, row_data in enumerate(data, start_row):
            for col_num, value in enumerate(row_data, 1):
                cell = self.worksheet.cell(row=row_num, column=col_num)
                cell.value = value
                cell.alignment = cell_alignment
                cell.border = border
                
                # رنگ ردیف‌های زوج
                if row_num % 2 == 0:
                    cell.fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    
    def auto_adjust_columns(self):
        """تنظیم خودکار عرض ستون‌ها"""
        if not self.worksheet:
            return
        
        for column in self.worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            self.worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def save(self, filename: str) -> bool:
        """ذخیره فایل Excel"""
        try:
            if not self.workbook:
                return False
            
            self.workbook.save(filename)
            return True
        except Exception as e:
            print(f"خطا در ذخیره فایل: {e}")
            return False
    
    def export_customers(self, customers: List[Dict], filename: str = "customers.xlsx") -> bool:
        """صادرات لیست مشتریان"""
        self.create_workbook("مشتریان")
        
        headers = ["شناسه", "نام", "تلفن", "ایمیل", "تاریخ تولد", "امتیاز", "یادداشت"]
        self.set_headers(headers)
        
        data = []
        for customer in customers:
            data.append([
                customer.get("id", ""),
                customer.get("name", ""),
                customer.get("phone", ""),
                customer.get("email", ""),
                customer.get("birth_date", ""),
                customer.get("loyalty_points", 0),
                customer.get("notes", "")
            ])
        
        self.add_rows(data)
        self.auto_adjust_columns()
        
        return self.save(filename)
    
    def export_inventory(self, items: List[Dict], filename: str = "inventory.xlsx") -> bool:
        """صادرات انبار"""
        self.create_workbook("انبار")
        
        headers = ["کد", "نام", "نوع", "واحد", "موجودی", "قیمت واحد", "ارزش کل"]
        self.set_headers(headers)
        
        data = []
        for item in items:
            quantity = item.get("quantity", 0)
            unit_price = item.get("unit_price", 0)
            total_value = quantity * unit_price
            
            data.append([
                item.get("code", ""),
                item.get("name", ""),
                item.get("item_type", ""),
                item.get("unit", ""),
                quantity,
                f"{unit_price:,}",
                f"{total_value:,}"
            ])
        
        self.add_rows(data)
        self.auto_adjust_columns()
        
        return self.save(filename)
    
    def export_invoices(self, invoices: List[Dict], filename: str = "invoices.xlsx") -> bool:
        """صادرات فاکتورها"""
        self.create_workbook("فاکتورها")
        
        headers = ["شماره فاکتور", "مشتری", "نوع", "مبلغ", "تخفیف", "مبلغ نهایی", "روش پرداخت", "تاریخ"]
        self.set_headers(headers)
        
        data = []
        for invoice in invoices:
            data.append([
                invoice.get("invoice_number", ""),
                invoice.get("customer_name", ""),
                invoice.get("invoice_type", ""),
                f"{invoice.get('subtotal', 0):,}",
                f"{invoice.get('discount_amount', 0):,}",
                f"{invoice.get('total_amount', 0):,}",
                invoice.get("payment_method", ""),
                invoice.get("created_at", "")
            ])
        
        self.add_rows(data)
        self.auto_adjust_columns()
        
        return self.save(filename)
    
    def export_sales_report(self, report_data: Dict, filename: str = "sales_report.xlsx") -> bool:
        """صادرات گزارش فروش"""
        self.create_workbook("گزارش فروش")
        
        # بخش خلاصه
        self.worksheet.cell(row=1, column=1).value = "گزارش فروش"
        self.worksheet.cell(row=1, column=1).font = Font(bold=True, size=16)
        
        self.worksheet.cell(row=2, column=1).value = f"تاریخ: {datetime.now().strftime('%Y/%m/%d')}"
        
        # آمار کلی
        row = 4
        stats = [
            ("فروش کل", f"{report_data.get('total_sales', 0):,} ریال"),
            ("تعداد فاکتور", report_data.get('invoice_count', 0)),
            ("متوسط فاکتور", f"{report_data.get('average_invoice', 0):,} ریال"),
            ("فروش کافه", f"{report_data.get('cafe_sales', 0):,} ریال"),
            ("فروش آرایشگاه", f"{report_data.get('barbershop_sales', 0):,} ریال"),
        ]
        
        for label, value in stats:
            self.worksheet.cell(row=row, column=1).value = label
            self.worksheet.cell(row=row, column=2).value = value
            row += 1
        
        # جدول فروش روزانه
        if "daily_sales" in report_data:
            row += 2
            self.worksheet.cell(row=row, column=1).value = "فروش روزانه"
            self.worksheet.cell(row=row, column=1).font = Font(bold=True, size=14)
            
            row += 1
            headers = ["تاریخ", "فروش", "تعداد فاکتور"]
            for col_num, header in enumerate(headers, 1):
                cell = self.worksheet.cell(row=row, column=col_num)
                cell.value = header
                cell.font = Font(bold=True)
            
            row += 1
            for daily in report_data["daily_sales"]:
                self.worksheet.cell(row=row, column=1).value = daily["date"]
                self.worksheet.cell(row=row, column=2).value = f"{daily['sales']:,}"
                self.worksheet.cell(row=row, column=3).value = daily["count"]
                row += 1
        
        self.auto_adjust_columns()
        
        return self.save(filename)
    
    def export_profit_report(self, report_data: Dict, filename: str = "profit_report.xlsx") -> bool:
        """صادرات گزارش سود"""
        self.create_workbook("گزارش سود")
        
        self.worksheet.cell(row=1, column=1).value = "گزارش سود خالص"
        self.worksheet.cell(row=1, column=1).font = Font(bold=True, size=16)
        
        row = 3
        items = [
            ("فروش کل", report_data.get('total_sales', 0), "10B981"),
            ("هزینه مواد مصرفی", -report_data.get('material_cost', 0), "F59E0B"),
            ("کمیسیون آرایشگران", -report_data.get('commission', 0), "F59E0B"),
            ("هزینه‌های جاری", -report_data.get('expenses', 0), "F59E0B"),
            ("سود خالص", report_data.get('net_profit', 0), "6366F1"),
        ]
        
        for label, value, color in items:
            self.worksheet.cell(row=row, column=1).value = label
            self.worksheet.cell(row=row, column=2).value = f"{abs(value):,} ریال"
            
            if label == "سود خالص":
                self.worksheet.cell(row=row, column=1).font = Font(bold=True)
                self.worksheet.cell(row=row, column=2).font = Font(bold=True)
            
            row += 1
        
        self.auto_adjust_columns()
        
        return self.save(filename)
